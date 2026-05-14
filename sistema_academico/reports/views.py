import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from django.views.generic import TemplateView
from django.db.models import Avg

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

from academic.models import Nota, Asistencia, Curso, Materia
from accounts.mixins import DocenteOAdminMixin


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_date(value, fallback=None):
    """Convierte string 'YYYY-MM-DD' a date, o devuelve fallback."""
    if value:
        try:
            return datetime.date.fromisoformat(value)
        except ValueError:
            pass
    return fallback


# ── Página de reportes (con filtros) ─────────────────────────────────────────

class ReportesView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/reportes.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from accounts.models import CustomUser
        ctx['estudiantes'] = CustomUser.objects.filter(
            rol='estudiante', activo=True
        ).order_by('last_name', 'first_name')
        ctx['cursos']   = Curso.objects.all().order_by('nombre')
        ctx['materias'] = Materia.objects.filter(activa=True).order_by('nombre')
        return ctx


# ── Boletín individual PDF ────────────────────────────────────────────────────

class BoletinPDFView(LoginRequiredMixin, View):
    """PDF con todas las notas de un estudiante, con filtro opcional por fechas."""

    def get(self, request, estudiante_id):
        from accounts.models import CustomUser
        from django.shortcuts import get_object_or_404
        estudiante = get_object_or_404(CustomUser, pk=estudiante_id)

        fecha_desde = _parse_date(request.GET.get('desde'))
        fecha_hasta = _parse_date(request.GET.get('hasta'), datetime.date.today())

        notas = Nota.objects.filter(
            estudiante=estudiante
        ).select_related('materia', 'curso').order_by('materia__nombre', 'fecha')

        if fecha_desde:
            notas = notas.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            notas = notas.filter(fecha__lte=fecha_hasta)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="boletin_{estudiante.username}.pdf"'
        )

        doc = SimpleDocTemplate(response, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        PURPLE = colors.HexColor('#5B21B6')
        LIGHT  = colors.HexColor('#F3F0FF')

        title_style = ParagraphStyle('title', parent=styles['Title'],
                                     textColor=PURPLE, fontSize=16)
        sub_style   = ParagraphStyle('sub', parent=styles['Normal'],
                                     textColor=colors.HexColor('#6B7280'), fontSize=9)

        elements = [
            Paragraph('Sistema de Gestión Académica', sub_style),
            Paragraph(f'Boletín de Notas — {estudiante.get_full_name()}', title_style),
            Spacer(1, 0.3*cm),
            Paragraph(
                f'Generado: {datetime.date.today().strftime("%d/%m/%Y")}' +
                (f'  |  Desde: {fecha_desde.strftime("%d/%m/%Y")}' if fecha_desde else '') +
                f'  |  Hasta: {fecha_hasta.strftime("%d/%m/%Y")}',
                sub_style
            ),
            Spacer(1, 0.5*cm),
        ]

        if notas.exists():
            tabla_data = [['Materia', 'Curso', 'Tipo', 'Nota', 'Fecha']]
            for n in notas:
                tabla_data.append([
                    n.materia.nombre,
                    n.curso.nombre,
                    n.get_tipo_display(),
                    str(n.valor),
                    n.fecha.strftime('%d/%m/%Y'),
                ])

            promedio = notas.aggregate(p=Avg('valor'))['p']
            tabla_data.append(['', '', 'PROMEDIO GENERAL', f'{promedio:.2f}', ''])

            tabla = Table(tabla_data, colWidths=[5*cm, 4*cm, 3.5*cm, 2*cm, 3*cm])
            tabla.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0),  (-1, 0),  PURPLE),
                ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
                ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0),  (-1, 0),  9),
                ('ROWBACKGROUNDS',(0, 1),  (-1, -2), [colors.white, LIGHT]),
                ('BACKGROUND',    (0, -1), (-1, -1), colors.HexColor('#EDE9FE')),
                ('FONTNAME',      (2, -1), (3, -1),  'Helvetica-Bold'),
                ('TEXTCOLOR',     (3, -1), (3, -1),  PURPLE),
                ('GRID',          (0, 0),  (-1, -1), 0.4, colors.HexColor('#DDD6FE')),
                ('FONTSIZE',      (0, 1),  (-1, -1), 9),
                ('ALIGN',         (3, 0),  (3, -1),  'CENTER'),
                ('ALIGN',         (4, 0),  (4, -1),  'CENTER'),
                ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
                ('TOPPADDING',    (0, 0),  (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0),  (-1, -1), 5),
            ]))
            elements.append(tabla)
        else:
            elements.append(
                Paragraph('No hay notas registradas para el período seleccionado.', styles['Normal'])
            )

        doc.build(elements)
        return response


# ── Acta de curso PDF ─────────────────────────────────────────────────────────

class ActaCursoPDFView(DocenteOAdminMixin, View):
    """PDF con las notas de todos los estudiantes de un curso."""

    def get(self, request, curso_id):
        from django.shortcuts import get_object_or_404
        curso       = get_object_or_404(Curso, pk=curso_id)
        estudiantes = curso.estudiantes.all().order_by('last_name', 'first_name')
        materias    = list(curso.materias.all().order_by('nombre'))

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="acta_{curso.nombre}_{curso.periodo}.pdf"'
        )

        doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        PURPLE = colors.HexColor('#5B21B6')
        LIGHT  = colors.HexColor('#F3F0FF')

        title_style = ParagraphStyle('title', parent=styles['Title'],
                                     textColor=PURPLE, fontSize=16)
        sub_style   = ParagraphStyle('sub', parent=styles['Normal'],
                                     textColor=colors.HexColor('#6B7280'), fontSize=9)

        elements = [
            Paragraph('Sistema de Gestión Académica', sub_style),
            Paragraph(f'Acta de Notas — {curso.nombre}', title_style),
            Spacer(1, 0.2*cm),
            Paragraph(
                f'Período: {curso.periodo}  |  Año: {curso.año}  |  '
                f'Generado: {datetime.date.today().strftime("%d/%m/%Y")}',
                sub_style
            ),
            Spacer(1, 0.5*cm),
        ]

        if not estudiantes.exists():
            elements.append(
                Paragraph('No hay estudiantes inscritos en este curso.', styles['Normal'])
            )
        else:
            encabezado = ['Estudiante'] + [m.nombre for m in materias] + ['Promedio']
            tabla_data = [encabezado]

            for est in estudiantes:
                fila = [est.get_full_name()]
                promedios_est = []
                for mat in materias:
                    prom = Nota.objects.filter(
                        estudiante=est, materia=mat, curso=curso
                    ).aggregate(p=Avg('valor'))['p']
                    fila.append(f'{prom:.1f}' if prom is not None else '—')
                    if prom is not None:
                        promedios_est.append(prom)
                prom_general = (sum(promedios_est) / len(promedios_est)) if promedios_est else None
                fila.append(f'{prom_general:.1f}' if prom_general is not None else '—')
                tabla_data.append(fila)

            col_w = [5*cm] + [3*cm] * len(materias) + [2.5*cm]
            tabla = Table(tabla_data, colWidths=col_w)
            tabla.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0),  (-1, 0),  PURPLE),
                ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
                ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0),  (-1, 0),  8),
                ('ROWBACKGROUNDS',(0, 1),  (-1, -1), [colors.white, LIGHT]),
                ('GRID',          (0, 0),  (-1, -1), 0.4, colors.HexColor('#DDD6FE')),
                ('FONTSIZE',      (0, 1),  (-1, -1), 8),
                ('ALIGN',         (1, 0),  (-1, -1), 'CENTER'),
                ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
                ('TOPPADDING',    (0, 0),  (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0),  (-1, -1), 5),
                ('FONTNAME',      (-1, 1), (-1, -1), 'Helvetica-Bold'),
                ('TEXTCOLOR',     (-1, 1), (-1, -1), PURPLE),
            ]))
            elements.append(tabla)

        doc.build(elements)
        return response


# ── Reporte Excel con filtros ─────────────────────────────────────────────────

class ReporteExcelView(LoginRequiredMixin, View):
    """Excel con notas y asistencia, filtradas por fecha, materia y/o curso."""

    def get(self, request):
        fecha_desde = _parse_date(request.GET.get('desde'))
        fecha_hasta = _parse_date(request.GET.get('hasta'), datetime.date.today())
        materia_id  = request.GET.get('materia')
        curso_id    = request.GET.get('curso')

        notas = Nota.objects.all().select_related('estudiante', 'materia', 'curso')
        if fecha_desde:
            notas = notas.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            notas = notas.filter(fecha__lte=fecha_hasta)
        if materia_id:
            notas = notas.filter(materia_id=materia_id)
        if curso_id:
            notas = notas.filter(curso_id=curso_id)
        notas = notas.order_by('curso__nombre', 'estudiante__last_name', 'materia__nombre')

        wb = openpyxl.Workbook()

        # ── Hoja 1: Notas ──────────────────────────────────────────────────
        ws = wb.active
        ws.title = 'Notas'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='5B21B6')
        alt_fill    = PatternFill('solid', fgColor='F3F0FF')
        center      = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A1:F1')
        ws['A1'] = 'Reporte de Notas — Sistema de Gestión Académica'
        ws['A1'].font      = Font(bold=True, size=13, color='5B21B6')
        ws['A1'].alignment = center

        ws.merge_cells('A2:F2')
        filtro_txt = f'Generado: {datetime.date.today().strftime("%d/%m/%Y")}'
        if fecha_desde:
            filtro_txt += f'  |  Desde: {fecha_desde.strftime("%d/%m/%Y")}'
        filtro_txt += f'  |  Hasta: {fecha_hasta.strftime("%d/%m/%Y")}'
        ws['A2'] = filtro_txt
        ws['A2'].font      = Font(italic=True, size=9, color='6B7280')
        ws['A2'].alignment = center

        encabezados = ['Estudiante', 'Materia', 'Curso', 'Tipo', 'Nota', 'Fecha']
        for col, enc in enumerate(encabezados, 1):
            cell = ws.cell(row=3, column=col, value=enc)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center

        for i, nota in enumerate(notas, 4):
            fila = [
                nota.estudiante.get_full_name(),
                nota.materia.nombre,
                nota.curso.nombre,
                nota.get_tipo_display(),
                float(nota.valor),
                nota.fecha.strftime('%d/%m/%Y'),
            ]
            for col, val in enumerate(fila, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.alignment = Alignment(vertical='center')
                if i % 2 == 0:
                    cell.fill = alt_fill
            nota_cell = ws.cell(row=i, column=5)
            nota_cell.font = Font(
                bold=True,
                color='DC2626' if nota.valor < 3 else '059669'
            )

        for col_letter, width in [('A', 28), ('B', 22), ('C', 20), ('D', 14), ('E', 8), ('F', 14)]:
            ws.column_dimensions[col_letter].width = width

        # ── Hoja 2: Asistencia ─────────────────────────────────────────────
        ws2 = wb.create_sheet('Asistencia')

        asistencias = Asistencia.objects.all().select_related(
            'estudiante', 'materia', 'curso'
        ).order_by('fecha', 'estudiante__last_name')

        if fecha_desde:
            asistencias = asistencias.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            asistencias = asistencias.filter(fecha__lte=fecha_hasta)
        if curso_id:
            asistencias = asistencias.filter(curso_id=curso_id)

        enc2 = ['Estudiante', 'Materia', 'Curso', 'Fecha', 'Presente']
        for col, enc in enumerate(enc2, 1):
            cell = ws2.cell(row=1, column=col, value=enc)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center

        for i, a in enumerate(asistencias, 2):
            fila = [
                a.estudiante.get_full_name(),
                a.materia.nombre,
                a.curso.nombre,
                a.fecha.strftime('%d/%m/%Y'),
                'Sí' if a.presente else 'No',
            ]
            for col, val in enumerate(fila, 1):
                cell = ws2.cell(row=i, column=col, value=val)
                cell.alignment = Alignment(vertical='center')
                if i % 2 == 0:
                    cell.fill = alt_fill

        for col_letter, width in [('A', 28), ('B', 22), ('C', 20), ('D', 14), ('E', 8)]:
            ws2.column_dimensions[col_letter].width = width

        # ── Respuesta ──────────────────────────────────────────────────────
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        nombre = f'reporte_notas_{datetime.date.today().isoformat()}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{nombre}"'
        wb.save(response)
        return response
